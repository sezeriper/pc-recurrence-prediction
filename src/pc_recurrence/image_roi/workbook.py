from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXPECTED_HEADERS: tuple[str, ...] = (
    "hasta id",
    "hasta no",
    "ameliyat şekli",
    "patoloji",
    "nüks",
    "yaş",
    "CA 19-9",
    "total bilirubin",
    "direkt bilirubin",
    "semptom",
    "serum albumin",
    "mutlak lenfosit",
    "CRP",
    "CALLY",
    "PNI",
    "Görüntü alanı",
    "BT raporu",
)


@dataclass(frozen=True)
class ImageWorkbookRow:
    patient_id: str
    row_number: int
    hasta_no: str | float | None
    dicom_folder: str | None
    image_range_raw: str | None


def _clean_scalar(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _dicom_folder(hasta_no: Any) -> str | None:
    """Derive the anonymized DICOM folder name from the workbook `hasta no`."""
    value = _clean_scalar(hasta_no)
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return f"PATIENT{text}" if text else None


def load_image_workbook(
    workbook_path: str | Path, sheet_name: str = "Sayfa1"
) -> list[ImageWorkbookRow]:
    path = Path(workbook_path)
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Expected worksheet {sheet_name!r}; found {workbook.sheetnames}")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(min_col=1, max_col=len(EXPECTED_HEADERS), values_only=True)
        raw_headers = next(rows, None)
        if raw_headers is None:
            raise ValueError("Workbook is empty")
        headers = tuple(
            str(value).strip() if value is not None else "" for value in raw_headers
        )
        if headers != EXPECTED_HEADERS:
            raise ValueError(
                f"Unexpected workbook schema. Expected {EXPECTED_HEADERS}; found {headers}"
            )

        records: list[ImageWorkbookRow] = []
        seen_ids: set[str] = set()
        for row_number, raw_row in enumerate(rows, start=2):
            if all(_clean_scalar(value) is None for value in raw_row):
                continue
            patient_id = _clean_scalar(raw_row[0])
            if not isinstance(patient_id, str) or not patient_id:
                raise ValueError(f"Row {row_number} has no valid patient ID")
            if patient_id in seen_ids:
                raise ValueError(f"Duplicate patient ID {patient_id!r} at row {row_number}")
            seen_ids.add(patient_id)

            hasta_no = _clean_scalar(raw_row[1])
            records.append(
                ImageWorkbookRow(
                    patient_id=patient_id,
                    row_number=row_number,
                    hasta_no=hasta_no,
                    dicom_folder=_dicom_folder(hasta_no),
                    image_range_raw=_clean_scalar(raw_row[15]),
                )
            )
    finally:
        workbook.close()

    if not records:
        raise ValueError("No populated patient rows were found")
    return records
